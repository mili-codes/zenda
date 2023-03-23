import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManage
import openpyxl


@pytest.fixture(scope="module")
def browser():
    # initialize the Chrome driver
    driver = webdriver.Chrome(ChromeDriverManager().install())
    yield driver
    # close the browser window
    driver.quit()


@pytest.fixture()
def test_data():
    # read test data from an Excel file
    workbook = openpyxl.load_workbook("./test_data/test_data_sheet.xlsx")
    worksheet = workbook.active
    # create a dictionary to store test data
    data = {"first_name": worksheet.cell(row=2, column=1).value,
            "last_name": worksheet.cell(row=2, column=2).value,
            "dob": worksheet.cell(row=2, column=3).value,
            "gender": worksheet.cell(row=2, column=4).value,
            "address": worksheet.cell(row=2, column=5).value}
    # return the test data dictionary
    return data